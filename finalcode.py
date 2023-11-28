import face_recognition
import numpy
import cv2
import os
from tkinter import *
import datetime
from pyzbar.pyzbar import decode
from openpyxl import workbook , load_workbook
# import pandas as pd
from PIL import Image,ImageTk
import time

# Creating App
root = Tk()
root.geometry("1920x1080")
root.configure(background="#ABB3B1")
var = StringVar()
var.set("Please show your QR")
title = Label(root, textvariable=var , font=("Arial",25) , background="#ABB3B1").place(x=620 , y=45)
L1 = Label(root)
L1.place(x=450,y=140)
# dece marking fucntion done
def markattendence(name):
    wb = load_workbook("Atten1.xlsx")
    ws = wb['Sheet1']
    dd = datetime.datetime.now()
    dd = dd.day
    dd = int(dd)
    i = 5;
    while ws.cell(row=i , column = 1).value != name:
        i = i+1
    # j = 2
    # while ws.cell(row=3 , column=j).value != dd:
    #     j = j+1
    ws.cell(row=i ,column = dd+1).value = "p"
    wb.save("Atten1.xlsx")
    
# taking faces and getting there encodings
def find_encodings():
    path = "C:/Users/lenovo/PycharmProjects/New Code/data"
    images = []
    classname = []
    mylist = os.listdir(path)
    for cl in mylist:
        curimg = cv2.imread(f'{path}/{cl}')
        curimg = cv2.cvtColor(curimg,cv2.COLOR_BGR2RGB)
        images.append(curimg)
        classname.append(os.path.splitext(cl)[0])
    # to update in excel
    wb = load_workbook("Atten1.xlsx")
    sheet = wb.active
    n=5
# sheet.cell(row=5 , column=2).value = 'p'
    for i in classname:
        c = 'A'+str(n)
        sheet[c] = i
        n = n+1
    wb.save("Atten1.xlsx")

    known_face_encodings  = []

    for img in images:
        currimg_encoding = face_recognition.face_encodings(img)[0]
        known_face_encodings.append(currimg_encoding)
    # encodings done
    # mapping data and names
    data = {}

    i = 0
    while(i < len(classname)):
        data[classname[i]] = known_face_encodings[i]
        i = i+1
    return data

data = find_encodings()
cap = cv2.VideoCapture(0)
while(True):
    # id = input("Enter the id :")
    # open camera and scanning qr
    var.set("Please show your QR")
    # print("Please Show you QR CODE")
    id = ''
    while (id == ''):
        suc, qr = cap.read()
        q = decode(qr)
        img = cv2.cvtColor(qr,cv2.COLOR_BGR2RGB)
        img = ImageTk.PhotoImage(Image.fromarray(img))
        L1['image'] = img
        root.update()
        # cv2.imshow('Webcam',qr)
        # cv2.waitKey(1)
        for i in q:
            id = i.data.decode()
    # capturing qr is done
    # scanning faces
    # print("Please look into the camera :")
    unknown_face_encoding = []
    while(len(unknown_face_encoding) == 0):
        #showing video and taking 1 pic from the video
        i = 0
        flag = 0
        while(i<150):
            i = i+1
            ran , img = cap.read()
            #resing the img
            imgs  = cv2.resize(img,(0,0),None,0.25,0.25);
            #taking a picture
            if(i == 75):
                img1 = imgs
            facloc = face_recognition.face_locations(imgs)
            # print(facloc," ",len(facloc))
            if(len(facloc) != 0):
                y1,x2,y2,x1 = facloc[0]
                y1,x2,y2,x1 = y1*4,x2*4,y2*4,x1*4
                cv2.rectangle(img,(x1,y1),(x2,y2),(0,255,0),2)
            var.set("Show your face")
            aimg = img 
            aimg = cv2.cvtColor(aimg,cv2.COLOR_BGR2RGB)
            aimg = ImageTk.PhotoImage(Image.fromarray(aimg))
            L1['image'] = aimg
            root.update()
        # take a pic
        # success , img = cap.read()
        # resize the pic
        # img = cv2.resize(img,(0,0),None,0.25,0.25)
        # convert to rgb
        img = cv2.cvtColor(img1,cv2.COLOR_BGR2RGB)
        unknown_face_encoding = face_recognition.face_encodings(img,model="cnn")
        if(len(unknown_face_encoding) == 0):
            var.set("Please Try again")
            root.update()
            time.sleep(3)
            # print('Please try again')
            flag = 1
            break
    if(flag == 1):
        continue
    try:
        matchrate = face_recognition.compare_faces(data[id],unknown_face_encoding)
        facedist = face_recognition.face_distance(data[id],unknown_face_encoding)
    except KeyError:
        var.set("Invalid")
        root.update()
        time.sleep(3)
        # print('Invalid ID')
        continue
    if(matchrate[0] == True and facedist < 0.5):
        var.set("Attendence Marked")
        img = cv2.imread("tick2.png")
        img = cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
        img = ImageTk.PhotoImage(Image.fromarray(img))
        L1['image'] = img
        # L1.place(x=500,y=190)
        root.update()
        markattendence(id)
        time.sleep(2)
        # print('Attendence Marked')
    else:
        var.set("Face not Matched")
        root.update()
        time.sleep(2)
        # print('Face not matched')
    # print(facedist)
    # print(matchrate)
    # root.mainloop()